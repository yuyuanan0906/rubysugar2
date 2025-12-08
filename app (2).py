# -*- coding: utf-8 -*-
"""
Created on Mon Dec  8 14:54:43 2025

@author: lianc
"""

import streamlit as st
import pandas as pd
from datetime import date, datetime
from fuzzywuzzy import fuzz
from openpyxl import load_workbook, Workbook
import os

FOOD_FILE = "foodssugar.xlsx"
RECORD_FILE = "Ruby_records.xlsx"


# ---------- 初始化 Excel ----------

def init_food_file():
    if not os.path.exists(FOOD_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "食物資料"
        ws.append(["食物名稱", "單位", "碳水化合物", "備註"])
        wb.save(FOOD_FILE)

def init_record_file():
    if not os.path.exists(RECORD_FILE):
        wb = Workbook()
        ws_food = wb.create_sheet("食物記錄")
        ws_food.append(["日期", "餐別", "食物名稱", "攝取量", "單位", "碳水化合物"])

        ws_insulin = wb.create_sheet("血糖與胰島素紀錄表")
        ws_insulin.append([
            "日期", "餐別", "總碳水量", "目前血糖值", "期望血糖值",
            "C/I值", "ISF值", "1C升高血糖", "碳水劑量", "矯正劑量",
            "總胰島素劑量", "餐後血糖值", "建議C/I值"
        ])
        wb.save(RECORD_FILE)


# ---------- Cache 讀檔 ----------

@st.cache_data
def load_foods_df() -> pd.DataFrame:
    init_food_file()
    df = pd.read_excel(FOOD_FILE, sheet_name="食物資料")
    return df

@st.cache_data
def load_records_df() -> pd.DataFrame:
    init_record_file()
    # 這裡僅示範讀取血糖紀錄，如需要可以再加食物紀錄
    df = pd.read_excel(RECORD_FILE, sheet_name="血糖與胰島素紀錄表")
    return df


# ---------- 寫檔（不 cache） ----------

def save_foods_df(df: pd.DataFrame):
    df.to_excel(FOOD_FILE, sheet_name="食物資料", index=False)
    load_foods_df.clear()   # 清除 cache，下次會重讀最新資料

def append_record(
    date_str, meal, calc_items, total_carb,
    current_glucose, target_glucose,
    ci, isf, c_raise,
    insulin_carb, insulin_corr, total_insulin
):
    init_record_file()
    wb = load_workbook(RECORD_FILE)

    ws_food = wb["食物記錄"]
    for item in calc_items:
        ws_food.append([
            date_str, meal,
            item["name"], item["amount"], item["unit"], item["carb"]
        ])
    ws_food.append(["", "", "", "", "總碳水", total_carb])

    ws_insulin = wb["血糖與胰島素紀錄表"]
    ws_insulin.append([
        date_str, meal, total_carb, current_glucose, target_glucose,
        ci, isf, c_raise, insulin_carb, insulin_corr, total_insulin,
        None,  # 餐後血糖值
        None   # 建議 C/I
    ])

    wb.save(RECORD_FILE)
    load_records_df.clear()


# ---------- 邏輯：找相似食物 / 計算 ----------

def find_similar_foods(df_foods: pd.DataFrame, keyword: str, threshold=60):
    if not keyword:
        return df_foods
    mask = df_foods["食物名稱"].apply(
        lambda name: fuzz.partial_ratio(str(keyword), str(name)) >= threshold
    )
    return df_foods[mask]

def round_insulin(value: float) -> float:
    decimal = value - int(value)
    if decimal <= 0.25:
        return round(int(value) + 0.0, 1)
    elif decimal <= 0.75:
        return round(int(value) + 0.5, 1)
    else:
        return round(int(value) + 1.0, 1)

def calc_insulin_dose(total_carb, ci, isf, current_glucose, target_glucose):
    insulin_carb = total_carb / ci if ci > 0 else 0
    insulin_corr = (current_glucose - target_glucose) / isf if isf > 0 else 0

    insulin_carb = round_insulin(insulin_carb)
    insulin_corr = round_insulin(insulin_corr)
    total_insulin = round_insulin(insulin_carb + insulin_corr)

    return insulin_carb, insulin_corr, total_insulin


# ---------- Streamlit App ----------

st.set_page_config(page_title="食物碳水與胰島素紀錄", layout="centered")

st.title("🍚 食物碳水與胰島素紀錄（手機版友善）")

# 用 session_state 存「這一餐的食物列表」
if "calc_items" not in st.session_state:
    st.session_state.calc_items = []

foods_df = load_foods_df()

st.markdown("### Step 1：設定日期與餐別")
col1, col2 = st.columns(2)
with col1:
    meal_date = st.date_input("日期", value=date.today())
with col2:
    meal = st.selectbox("餐別", ["早餐", "午餐", "晚餐", "宵夜"])

st.divider()

st.markdown("### Step 2：加入本餐食物")

with st.form("add_food_form", clear_on_submit=True):
    keyword = st.text_input("🔍 搜尋食物名稱（關鍵字）")
    filtered = find_similar_foods(foods_df, keyword)

    if filtered.empty:
        st.info("查無相似食物，可以到『食物管理頁』新增。")
        selected_food = None
    else:
        food_options = filtered["食物名稱"] + "｜每" + filtered["單位"] + f" 含 " + filtered["碳水化合物"].astype(str) + "g"
        idx = st.selectbox("選擇食物", range(len(filtered)), format_func=lambda i: food_options.iloc[i])
        selected_row = filtered.iloc[idx]
        selected_food = {
            "name": selected_row["食物名稱"],
            "unit": selected_row["單位"],
            "carb_per_unit": float(selected_row["碳水化合物"]),
        }

    amount = st.number_input("攝取量（同上單位）", min_value=0.0, step=1.0)

    submitted = st.form_submit_button("➕ 加入本餐")

    if submitted:
        if (not selected_food) or amount <= 0:
            st.warning("請先選食物並輸入大於 0 的攝取量")
        else:
            carb = round(selected_food["carb_per_unit"] * amount, 2)
            st.session_state.calc_items.append({
                "name": selected_food["name"],
                "unit": selected_food["unit"],
                "amount": amount,
                "carb": carb,
            })
            st.success(f"已加入：{selected_food['name']}，碳水 {carb} g")

# 顯示目前本餐食物清單
if st.session_state.calc_items:
    st.markdown("#### 本餐食物清單")
    df_current = pd.DataFrame(st.session_state.calc_items)
    df_current_display = df_current.rename(columns={
        "name": "食物名稱",
        "unit": "單位",
        "amount": "攝取量",
        "carb": "碳水(g)"
    })
    st.dataframe(df_current_display, use_container_width=True)

    total_carb = round(df_current["carb"].sum(), 2)
    st.subheader(f"本餐總碳水量：**{total_carb} g**")

    if st.button("🧹 清除本餐所有食物"):
        st.session_state.calc_items = []
        st.experimental_rerun()
else:
    total_carb = 0.0
    st.info("尚未加入任何食物。")

st.divider()

st.markdown("### Step 3：輸入血糖與參數，計算胰島素劑量")

with st.form("calc_insulin_form"):
    col1, col2 = st.columns(2)
    with col1:
        current_glucose = st.number_input("🩸 目前血糖值", min_value=0, step=1)
        target_glucose = st.number_input("🎯 期望血糖值", min_value=0, step=1, value=100)
    with col2:
        ci = st.number_input("C/I 值", min_value=0.0, step=0.1)
        isf = st.number_input("ISF 值", min_value=0.0, step=0.1, value=50.0)
    c_raise = st.number_input("1C 升高血糖", min_value=0.0, step=0.1, value=0.0)

    calc_and_save = st.form_submit_button("🧮 計算胰島素並儲存")

    if calc_and_save:
        if not st.session_state.calc_items:
            st.warning("尚未加入任何食物，本餐碳水為 0，仍可儲存血糖與參數。")

        if ci <= 0 or isf <= 0:
            st.error("請輸入有效的 C/I 與 ISF 值（需大於 0）")
        else:
            insulin_carb, insulin_corr, total_insulin = calc_insulin_dose(
                total_carb, ci, isf, current_glucose, target_glucose
            )

            st.markdown(f"""
            **計算結果：**

            - 碳水劑量：`{insulin_carb} U`  
            - 矯正劑量：`{insulin_corr} U`  
            - 總胰島素劑量：`{total_insulin} U`
            """)

            date_str = meal_date.strftime("%Y-%m-%d")
            append_record(
                date_str, meal,
                st.session_state.calc_items, total_carb,
                int(current_glucose), int(target_glucose),
                float(ci), float(isf), float(c_raise),
                float(insulin_carb), float(insulin_corr), float(total_insulin)
            )

            st.success(f"已儲存 {date_str} {meal} 的紀錄")
            # 儲存一餐後，清除本餐食物
            st.session_state.calc_items = []
